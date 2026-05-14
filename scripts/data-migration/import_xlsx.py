#!/usr/bin/env python3
"""
Import historical asset snapshots from an xlsx workbook into a Strata SQLite DB.

The xlsx layout assumed:
- A single sheet (default "Banques") with multiple "tables" stacked vertically.
- Each table starts with a header row whose first column is a year label
  (e.g. "2023/2024", "20-21") and subsequent columns are dates (mixed
  formats: datetime, "DD/MM/YYYY", or French long-form like "14 Mars 2017").
- Following rows are one per asset; cell values are EUR amounts at that date.
- Aggregate rows (TOTAL, GRAND TOTAL, Growth, etc.) are skipped via config.

All personal data (paths, row→asset mapping, rename rules) lives in a JSON
config file passed via --config. This script contains zero personal data and
is safe to commit. A sanitised template (`mapping.example.json`) is included.

Config schema:
{
  "xlsx_path": "/abs/path/to/workbook.xlsx",
  "sheet": "Banques",
  "db_path": "/abs/path/to/strata.db",
  "table_headers": ["2026/2027", "2023/2024", ...],   // first-cell strings marking new tables
  "skip_rows": ["TOTAL", "GRAND TOTAL", "Growth", ...],  // case-insensitive, prefix match
  "renames": [ {"from": "Old Name", "to": "New Name"} ],
  "rows": {
    "<xlsx row label (lowercase)>": {
      "action": "map" | "create" | "skip",
      "asset_name": "<target asset>",            // for map / create
      "asset_type_code": "CHECKING_ACCOUNT",     // for create only
      "context": "default" | "post_2024"         // optional disambiguator
    },
    ...
  },
  "context_overrides": [                          // optional row-level routing
    {
      "row_label": "ambiguous label",
      "before": "2024-01-01",
      "asset_name": "Asset Variant A"
    }
  ]
}

Usage:
    python3 import_xlsx.py --config scripts/data-migration/mapping.local.json [--dry-run]
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sqlite3
import sys
import uuid
from pathlib import Path
from typing import Any, Iterable

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover
    print("openpyxl is required. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# --- Date parsing ------------------------------------------------------------

FRENCH_MONTHS = {
    "janvier": 1, "février": 2, "fevrier": 2, "mars": 3, "avril": 4, "mai": 5,
    "juin": 6, "juillet": 7, "août": 8, "aout": 8, "septembre": 9,
    "octobre": 10, "novembre": 11, "décembre": 12, "decembre": 12,
}

DATE_PATTERN_DMY = re.compile(r"^\s*(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{2,4})\s*$")
DATE_PATTERN_FRENCH = re.compile(
    r"^\s*(?:(\d{1,2})\s*)?([A-Za-zéûÉÛîôÎÔ]+)\s+(\d{4})\s*$"
)
COLUMN_PLACEHOLDER = re.compile(r"^\s*Column\s+\d+\s*$", re.IGNORECASE)


def parse_date(cell: Any) -> dt.date | None:
    """Parse a single header cell into a date. Returns None if cell is empty."""
    if cell is None:
        return None
    if isinstance(cell, dt.datetime):
        return cell.date()
    if isinstance(cell, dt.date):
        return cell
    if isinstance(cell, str):
        s = cell.strip()
        if not s:
            return None
        if COLUMN_PLACEHOLDER.match(s):
            return None
        m = DATE_PATTERN_DMY.match(s)
        if m:
            day, month, year = (int(x) for x in m.groups())
            if year < 100:
                year += 2000
            return dt.date(year, month, day)
        m = DATE_PATTERN_FRENCH.match(s)
        if m:
            day_str, month_str, year_str = m.groups()
            month_key = month_str.lower()
            if month_key in FRENCH_MONTHS:
                month = FRENCH_MONTHS[month_key]
                day = int(day_str) if day_str else 15
                return dt.date(int(year_str), month, day)
    return None


# --- Row classification ------------------------------------------------------

def is_table_header(label: str, table_headers: Iterable[str]) -> bool:
    if not label:
        return False
    norm = label.strip()
    return any(norm == th.strip() for th in table_headers)


def is_skip_row(label: str, skip_prefixes: Iterable[str]) -> bool:
    if not label:
        return True
    norm = label.strip().lower()
    return any(norm.startswith(p.strip().lower()) for p in skip_prefixes)


def normalise_row_label(label: str) -> str:
    return label.strip().lower()


# --- Config-driven asset resolution ------------------------------------------

def resolve_asset(
    row_label: str,
    observed_at: dt.date,
    config: dict,
) -> tuple[str, dict] | None:
    """
    Resolve a (row_label, date) tuple to an asset config entry.
    Returns (asset_name, row_config) or None if row should be skipped.
    Context overrides are consulted first (route same label by snapshot date).
    """
    norm = normalise_row_label(row_label)
    for ovr in config.get("context_overrides", []):
        if normalise_row_label(ovr.get("row_label", "")) == norm:
            before_str = ovr.get("before")
            after_str = ovr.get("after")
            if before_str and observed_at < dt.date.fromisoformat(before_str):
                return (ovr["asset_name"], {"action": "map"})
            if after_str and observed_at >= dt.date.fromisoformat(after_str):
                return (ovr["asset_name"], {"action": "map"})

    rows = config.get("rows", {})
    if norm in rows:
        row_cfg = rows[norm]
        if row_cfg.get("action") == "skip":
            return None
        return (row_cfg["asset_name"], row_cfg)
    return None


# --- DB helpers --------------------------------------------------------------

def fetch_asset_id(conn: sqlite3.Connection, name: str) -> str | None:
    cur = conn.execute('SELECT id FROM assets WHERE name = ?', (name,))
    row = cur.fetchone()
    return row[0] if row else None


def fetch_asset_type_id(conn: sqlite3.Connection, code: str) -> str:
    cur = conn.execute('SELECT id FROM asset_types WHERE code = ?', (code,))
    row = cur.fetchone()
    if not row:
        raise RuntimeError(f"Asset type with code='{code}' not found in DB.")
    return row[0]


def snapshot_exists(conn: sqlite3.Connection, asset_id: str, observed_at: dt.date) -> bool:
    iso = f"{observed_at.isoformat()} 00:00:00"
    iso_z = f"{observed_at.isoformat()}T00:00:00.000Z"
    cur = conn.execute(
        'SELECT 1 FROM asset_snapshots WHERE asset_id = ? AND observed_at IN (?, ?)',
        (asset_id, iso, iso_z),
    )
    return cur.fetchone() is not None


def format_decimal(value: float) -> str:
    return f"{round(float(value), 2):.2f}"


def now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def ensure_asset(
    conn: sqlite3.Connection,
    asset_name: str,
    asset_type_code: str,
    first_observed_at: dt.date,
    first_value: float,
) -> tuple[str, bool]:
    """Return (asset_id, was_created). Adds an ACQUIRE txn if asset is created."""
    existing = fetch_asset_id(conn, asset_name)
    if existing:
        return existing, False
    asset_type_id = fetch_asset_type_id(conn, asset_type_code)
    asset_id = str(uuid.uuid4())
    now = now_iso()
    conn.execute(
        'INSERT INTO assets (id, name, quantity, disposed, asset_type_id, created_at, updated_at) '
        'VALUES (?, ?, NULL, 0, ?, ?, ?)',
        (asset_id, asset_name, asset_type_id, now, now),
    )
    txn_id = str(uuid.uuid4())
    occurred_at = f"{first_observed_at.isoformat()} 00:00:00"
    conn.execute(
        'INSERT INTO transactions (id, asset_id, type, unit_price, quantity, currency, occurred_at, created_at) '
        'VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
        (txn_id, asset_id, "ACQUIRE", format_decimal(first_value), "1", "EUR", occurred_at, now),
    )
    return asset_id, True


def insert_snapshot(
    conn: sqlite3.Connection,
    asset_id: str,
    observed_at: dt.date,
    value: float,
) -> bool:
    """Insert (asset_id, observed_at, value). Returns True if a row was inserted."""
    if snapshot_exists(conn, asset_id, observed_at):
        return False
    snap_id = str(uuid.uuid4())
    observed = f"{observed_at.isoformat()} 00:00:00"
    conn.execute(
        'INSERT INTO asset_snapshots (id, asset_id, value, observed_at, created_at) '
        'VALUES (?, ?, ?, ?, ?)',
        (snap_id, asset_id, format_decimal(value), observed, now_iso()),
    )
    return True


def apply_renames(conn: sqlite3.Connection, renames: list[dict]) -> list[tuple[str, str, bool]]:
    """Apply name updates. Returns list of (from, to, applied)."""
    result = []
    for r in renames:
        old, new = r["from"], r["to"]
        cur = conn.execute('SELECT id FROM assets WHERE name = ?', (old,))
        row = cur.fetchone()
        if row:
            conn.execute('UPDATE assets SET name = ?, updated_at = ? WHERE id = ?', (new, now_iso(), row[0]))
            result.append((old, new, True))
        else:
            result.append((old, new, False))
    return result


# --- Sheet walker ------------------------------------------------------------

def collect_assets_first_snapshot(
    rows: list[tuple[Any, ...]],
    config: dict,
) -> dict[str, tuple[dt.date, float]]:
    """First pass: for each create-action asset, find its earliest (date, value)."""
    earliest: dict[str, tuple[dt.date, float]] = {}
    current_dates: list[dt.date | None] = []
    table_headers = config.get("table_headers", [])
    skip_prefixes = config.get("skip_rows", [])

    for row in rows:
        if not row:
            continue
        first = row[0]
        if isinstance(first, str) and is_table_header(first, table_headers):
            current_dates = [parse_date(c) for c in row[1:]]
            continue
        if not isinstance(first, str) or is_skip_row(first, skip_prefixes):
            continue
        for idx, val in enumerate(row[1:]):
            if val is None or idx >= len(current_dates):
                continue
            observed = current_dates[idx]
            if observed is None:
                continue
            resolved = resolve_asset(first, observed, config)
            if resolved is None:
                continue
            asset_name, row_cfg = resolved
            if row_cfg.get("action") != "create":
                continue
            try:
                fval = float(val)
            except (TypeError, ValueError):
                continue
            if asset_name not in earliest or observed < earliest[asset_name][0]:
                earliest[asset_name] = (observed, fval)
    return earliest


def walk_and_import(
    conn: sqlite3.Connection,
    rows: list[tuple[Any, ...]],
    config: dict,
    dry_run: bool,
) -> dict:
    """Second pass: actually emit snapshots."""
    report = {
        "snapshots_inserted": 0,
        "snapshots_skipped_existing": 0,
        "snapshots_skipped_unresolved": 0,
        "rows_skipped_aggregate": 0,
        "unparseable_dates": [],
        "per_asset": {},
        "created_assets": [],
        "renames": [],
        "unresolved_labels": set(),
    }
    table_headers = config.get("table_headers", [])
    skip_prefixes = config.get("skip_rows", [])

    # Phase A: renames
    rename_results = apply_renames(conn, config.get("renames", []))
    report["renames"] = rename_results

    # Phase B: pre-compute first snapshot per create-asset (for ACQUIRE txn dating)
    earliest = collect_assets_first_snapshot(rows, config)
    for asset_name, (date, value) in earliest.items():
        rows_cfg = config.get("rows", {})
        row_cfg = next(
            (rows_cfg[k] for k in rows_cfg if rows_cfg[k].get("asset_name") == asset_name and rows_cfg[k].get("action") == "create"),
            None,
        )
        if row_cfg is None:
            continue
        asset_id, was_created = ensure_asset(
            conn, asset_name, row_cfg["asset_type_code"], date, value
        )
        if was_created:
            report["created_assets"].append({"name": asset_name, "first_date": date.isoformat(), "first_value": value})

    # Phase C: walk and emit snapshots
    current_dates: list[dt.date | None] = []
    for row in rows:
        if not row:
            continue
        first = row[0]
        if isinstance(first, str) and is_table_header(first, table_headers):
            current_dates = [parse_date(c) for c in row[1:]]
            for raw, parsed in zip(row[1:], current_dates):
                if raw not in (None, "") and parsed is None and not (isinstance(raw, str) and COLUMN_PLACEHOLDER.match(raw)):
                    report["unparseable_dates"].append(repr(raw))
            continue
        if not isinstance(first, str):
            continue
        if is_skip_row(first, skip_prefixes):
            report["rows_skipped_aggregate"] += 1
            continue
        for idx, val in enumerate(row[1:]):
            if val is None or idx >= len(current_dates):
                continue
            observed = current_dates[idx]
            if observed is None:
                continue
            try:
                fval = float(val)
            except (TypeError, ValueError):
                continue
            resolved = resolve_asset(first, observed, config)
            if resolved is None:
                report["snapshots_skipped_unresolved"] += 1
                report["unresolved_labels"].add(first.strip())
                continue
            asset_name, _row_cfg = resolved
            asset_id = fetch_asset_id(conn, asset_name)
            if asset_id is None:
                report["snapshots_skipped_unresolved"] += 1
                report["unresolved_labels"].add(f"NO-ASSET:{asset_name}")
                continue
            inserted = insert_snapshot(conn, asset_id, observed, fval)
            if inserted:
                report["snapshots_inserted"] += 1
                report["per_asset"][asset_name] = report["per_asset"].get(asset_name, 0) + 1
            else:
                report["snapshots_skipped_existing"] += 1

    report["unresolved_labels"] = sorted(report["unresolved_labels"])
    return report


def load_workbook_rows(xlsx_path: Path, sheet: str) -> list[tuple[Any, ...]]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb[sheet]
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def main() -> int:
    parser = argparse.ArgumentParser(description="Import xlsx snapshots into a Strata SQLite DB.")
    parser.add_argument("--config", required=True, help="Path to local mapping JSON config.")
    parser.add_argument("--dry-run", action="store_true", help="Roll back the transaction at the end.")
    args = parser.parse_args()

    config_path = Path(args.config).expanduser()
    if not config_path.exists():
        print(f"ERROR: config not found at {config_path}", file=sys.stderr)
        return 1
    with config_path.open(encoding="utf-8") as f:
        config = json.load(f)

    xlsx_path = Path(config["xlsx_path"]).expanduser()
    db_path = Path(config["db_path"]).expanduser()
    sheet = config.get("sheet", "Banques")
    if not xlsx_path.exists():
        print(f"ERROR: xlsx not found at {xlsx_path}", file=sys.stderr)
        return 1
    if not db_path.exists():
        print(f"ERROR: DB not found at {db_path}", file=sys.stderr)
        return 1

    rows = load_workbook_rows(xlsx_path, sheet)
    conn = sqlite3.connect(str(db_path))
    try:
        conn.execute("PRAGMA foreign_keys = ON")
        report = walk_and_import(conn, rows, config, dry_run=args.dry_run)
        if args.dry_run:
            conn.rollback()
            print("[DRY RUN] transaction rolled back.")
        else:
            conn.commit()
    finally:
        conn.close()

    print("\n=== Import report ===")
    print(f"Snapshots inserted          : {report['snapshots_inserted']}")
    print(f"Snapshots skipped (exists)  : {report['snapshots_skipped_existing']}")
    print(f"Snapshots skipped (unresol) : {report['snapshots_skipped_unresolved']}")
    print(f"Rows skipped (aggregate)    : {report['rows_skipped_aggregate']}")
    print(f"Renames applied             : {sum(1 for r in report['renames'] if r[2])}/{len(report['renames'])}")
    print(f"Assets created              : {len(report['created_assets'])}")
    if report["unparseable_dates"]:
        print(f"Unparseable dates           : {report['unparseable_dates']}")
    if report["unresolved_labels"]:
        print(f"Unresolved row labels       : {report['unresolved_labels']}")
    print("\nPer-asset snapshot counts:")
    for name, n in sorted(report["per_asset"].items(), key=lambda x: -x[1]):
        print(f"  {name:<32} {n}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
